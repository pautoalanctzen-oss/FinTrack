# -*- coding: utf-8 -*-
from fastapi import FastAPI, Request, Form, status
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
import uvicorn
from fastapi.middleware.cors import CORSMiddleware
from fastapi import HTTPException
from datetime import date, datetime
import sqlite3
import bcrypt
from contextlib import contextmanager
import os
import re
import logging
import sys
import traceback
from typing import Union
import time
from starlette.middleware.base import BaseHTTPMiddleware
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    POSTGRES_AVAILABLE = True
except ImportError:
    POSTGRES_AVAILABLE = False

# =======================
# CONFIGURACIÓN DE LOGGING
# =======================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('backend.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

app = FastAPI(title="Sistema de Login y Registro", version="1.0.0")

# Configuración de base de datos
DATABASE_URL = os.getenv("DATABASE_URL")  # PostgreSQL en producción (externo o interno)
USE_POSTGRES = bool(DATABASE_URL) and POSTGRES_AVAILABLE
DB_PATH = os.path.join(os.path.dirname(__file__), "users.db")  # SQLite local

if USE_POSTGRES:
    logger.info("Usando PostgreSQL en producción")
else:
    logger.info(f"Usando SQLite: {DB_PATH}")


@contextmanager
def get_db():
    """Context manager para manejar conexiones a la base de datos (PostgreSQL o SQLite)."""
    if USE_POSTGRES:
        # PostgreSQL en producción
        max_retries = 3
        retry_delay = 0.5
        conn = None
        
        for attempt in range(max_retries):
            try:
                conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
                conn.set_session(autocommit=False)
                break
            except Exception as e:
                logger.error(f"Error al conectar a PostgreSQL (intento {attempt + 1}/{max_retries}): {e}")
                if conn:
                    conn.close()
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                else:
                    logger.critical("No se pudo establecer conexión con PostgreSQL")
                    raise
        
        try:
            yield conn
        except Exception as e:
            logger.error(f"Error durante operación de DB: {e}")
            if conn:
                conn.rollback()
            raise
        finally:
            if conn:
                conn.close()
    else:
        # SQLite en desarrollo
        max_retries = 3
        retry_delay = 0.5
        conn = None
        
        for attempt in range(max_retries):
            try:
                conn = sqlite3.connect(DB_PATH, timeout=10.0)
                conn.row_factory = sqlite3.Row
                conn.execute("PRAGMA encoding = 'UTF-8'")
                conn.execute("SELECT 1")
                break
            except sqlite3.Error as e:
                logger.error(f"Error al conectar a SQLite (intento {attempt + 1}/{max_retries}): {e}")
                if conn:
                    conn.close()
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                else:
                    logger.critical("No se pudo establecer conexión con la base de datos")
                    raise
        
        try:
            yield conn
        except Exception as e:
            logger.error(f"Error durante operación de DB: {e}")
            if conn:
                conn.rollback()
            raise
        finally:
            if conn:
                conn.close()


def init_db():
    """Inicializa la base de datos con todas las tablas necesarias."""
    try:
        logger.info("Inicializando base de datos...")
        with get_db() as conn:
            cursor = conn.cursor()
            
            if USE_POSTGRES:
                # PostgreSQL - usar SERIAL en lugar de AUTOINCREMENT
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS users (
                        id SERIAL PRIMARY KEY,
                        email TEXT UNIQUE NOT NULL,
                        username TEXT UNIQUE NOT NULL,
                        birthdate TEXT NOT NULL,
                        password_hash TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS obras (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        ubicacion TEXT,
                        estado TEXT DEFAULT 'activa',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS clientes (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        cedula TEXT,
                        obra TEXT,
                        estado TEXT DEFAULT 'activo',
                        fecha TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS productos (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        precio REAL NOT NULL DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS registros (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL,
                        fecha TEXT,
                        obra TEXT,
                        totalCantidad INTEGER DEFAULT 0,
                        totalCobrar REAL DEFAULT 0,
                        totalPagado REAL DEFAULT 0,
                        status TEXT DEFAULT 'pendiente',
                        clientesAdicionales TEXT,
                        detalles TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
            else:
                # SQLite - usar AUTOINCREMENT
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        email TEXT UNIQUE NOT NULL,
                        username TEXT UNIQUE NOT NULL,
                        birthdate TEXT NOT NULL,
                        password_hash TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS obras (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        ubicacion TEXT,
                        estado TEXT DEFAULT 'activa',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS clientes (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        cedula TEXT,
                        obra TEXT,
                        estado TEXT DEFAULT 'activo',
                        fecha TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS productos (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        precio REAL NOT NULL DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS registros (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        fecha TEXT,
                        obra TEXT,
                        totalCantidad INTEGER DEFAULT 0,
                        totalCobrar REAL DEFAULT 0,
                        totalPagado REAL DEFAULT 0,
                        status TEXT DEFAULT 'pendiente',
                        clientesAdicionales TEXT,
                        detalles TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
            
            conn.commit()
        logger.info("Base de datos inicializada correctamente")
    except Exception as e:
        logger.warning(f"No se pudo inicializar todas las tablas en DB: {e}. Continuando...")
            
            if USE_POSTGRES:
                # PostgreSQL - usar SERIAL en lugar de AUTOINCREMENT
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS users (
                        id SERIAL PRIMARY KEY,
                        email TEXT UNIQUE NOT NULL,
                        username TEXT UNIQUE NOT NULL,
                        birthdate TEXT NOT NULL,
                        password_hash TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS obras (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        ubicacion TEXT,
                        estado TEXT DEFAULT 'activa',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS clientes (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        cedula TEXT,
                        obra TEXT,
                        estado TEXT DEFAULT 'activo',
                        fecha TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        fecha TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS productos (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        precio REAL NOT NULL DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS registros (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL,
                        fecha TEXT,
                        obra TEXT,
                        totalCantidad INTEGER DEFAULT 0,
                        totalCobrar REAL DEFAULT 0,
                        totalPagado REAL DEFAULT 0,
                        status TEXT DEFAULT 'pendiente',
                        clientesAdicionales TEXT,
                        detalles TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
            else:
                # SQLite - usar AUTOINCREMENT
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        email TEXT UNIQUE NOT NULL,
                        username TEXT UNIQUE NOT NULL,
                        birthdate TEXT NOT NULL,
                        password_hash TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS obras (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        ubicacion TEXT,
                        estado TEXT DEFAULT 'activa',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS clientes (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        cedula TEXT,
                        obra TEXT,
                        estado TEXT DEFAULT 'activo',
                        fecha TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS productos (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        nombre TEXT NOT NULL,
                        precio REAL NOT NULL DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS registros (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        fecha TEXT,
                        obra TEXT,
                        totalCantidad INTEGER DEFAULT 0,
                        totalCobrar REAL DEFAULT 0,
                        totalPagado REAL DEFAULT 0,
                        status TEXT DEFAULT 'pendiente',
                        clientesAdicionales TEXT,
                        detalles TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
            
            conn.commit()
        logger.info("Base de datos inicializada correctamente")
    except Exception as e:
        logger.warning(f"No se pudo inicializar todas las tablas en DB: {e}. Continuando con fallback...")


# Inicializar DB al arrancar
init_db()


def get_last_insert_id(cursor):
    """Obtiene el último ID insertado (compatible con SQLite y PostgreSQL)."""
    if USE_POSTGRES:
        return cursor.fetchone()[0]
    else:
        return cursor.lastrowid


def sql(query, params=None):
    """
    Convierte automáticamente placeholders SQL de SQLite (?) a PostgreSQL (%s).
    Uso: sql("SELECT * FROM users WHERE id = ?", (user_id,))
    """
    if USE_POSTGRES and query:
        # Convertir ? a %s para PostgreSQL
        query = query.replace("?", "%s")
    return (query, params) if params else (query,)


def ensure_demo_user():
    """Crea un usuario de prueba si no existe para facilitar el acceso.

    Usuario: demo
    Contraseña: Demo1234
    """
    try:
        email = "demo@example.com"
        username = "demo"
        birthdate = "2000-01-01"
        password_plain = "Demo1234"
        with get_db() as conn:
            cur = conn.execute("SELECT id FROM users WHERE username = ? OR email = ?", (username, email))
            row = cur.fetchone()
            password_hash = bcrypt.hashpw(password_plain.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

            if row:
                # Actualizar contraseña para asegurar acceso
                conn.execute(
                    "UPDATE users SET password_hash = ?, birthdate = ?, email = ? WHERE id = ?",
                    (password_hash, birthdate, email, row["id"]),
                )
                conn.commit()
                logger.info("Usuario demo ya existía; contraseña restablecida")
                return

            conn.execute(
                "INSERT INTO users (email, username, birthdate, password_hash) VALUES (?, ?, ?, ?)",
                (email, username, birthdate, password_hash),
            )
            conn.commit()
            logger.info("Usuario demo creado exitosamente")
    except Exception as e:
        logger.error(f"Error al crear usuario demo: {e}")
        # No es crítico, no detenemos la app


# Asegurar usuario demo persistente
ensure_demo_user()

# Montar archivos estáticos del frontend
frontend_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend")
if os.path.exists(frontend_path):
    app.mount("/frontend", StaticFiles(directory=frontend_path), name="frontend")
    # Montar la carpeta assets directamente para que sea accesible desde /assets
    assets_path = os.path.join(frontend_path, "assets")
    if os.path.exists(assets_path):
        app.mount("/assets", StaticFiles(directory=assets_path), name="assets")
    # Montar archivos JS directamente
    app.mount("/api.js", StaticFiles(directory=frontend_path, html=True), name="static_js")

# Directorio de plantillas (para compatibilidad, aunque ahora usamos archivos estáticos)
templates = Jinja2Templates(directory="templates")

# =======================
# MIDDLEWARE DE ERROR HANDLING GLOBAL
# =======================
@app.middleware("http")
async def error_handling_middleware(request: Request, call_next):
    """Middleware que captura todas las excepciones no manejadas."""
    try:
        return await call_next(request)
    except Exception as exc:
        logger.error(f"Error no manejado en {request.method} {request.url.path}: {exc}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"message": "Error interno del servidor. El error ha sido registrado."},
        )

# Middleware para UTF-8 y anti-caché
class UTF8Middleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        
        # Forzar UTF-8 en todas las respuestas de texto
        if 'content-type' in response.headers:
            content_type = response.headers['content-type']
            # JSON
            if 'application/json' in content_type and 'charset' not in content_type:
                response.headers['content-type'] = 'application/json; charset=utf-8'
            # HTML
            elif 'text/html' in content_type and 'charset' not in content_type:
                response.headers['content-type'] = 'text/html; charset=utf-8'
            # JavaScript
            elif 'javascript' in content_type and 'charset' not in content_type:
                response.headers['content-type'] = content_type + '; charset=utf-8'
        
        # Agregar cabeceras anti-caché para archivos HTML, CSS, JS e imágenes
        if any(request.url.path.endswith(ext) for ext in ['.html', '.css', '.js', '.png', '.jpg', '.svg', '.jpeg']) or request.url.path == '/':
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
        
        return response

app.add_middleware(UTF8Middleware)

# CORS para permitir peticiones desde el frontend (archivo local o localhost)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # en producción, especifica orígenes permitidos
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
async def home():
    """Sirve la página de login (index.html)."""
    index_path = os.path.join(frontend_path, "index.html")
    if os.path.exists(index_path):
        return FileResponse(
            index_path,
            media_type="text/html; charset=utf-8",
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Content-Type": "text/html; charset=utf-8"
            }
        )
    return {"message": "Bienvenido a la API"}


@app.get("/index.html")
async def index_html():
    """Sirve la página de login (index.html)."""
    index_path = os.path.join(frontend_path, "index.html")
    if os.path.exists(index_path):
        return FileResponse(
            index_path,
            media_type="text/html; charset=utf-8",
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Content-Type": "text/html; charset=utf-8"
            }
        )
    return {"detail": "Not Found"}


@app.get("/register.html")
async def register_html():
    """Sirve la página de registro."""
    register_path = os.path.join(frontend_path, "register.html")
    if os.path.exists(register_path):
        return FileResponse(
            register_path,
            media_type="text/html; charset=utf-8",
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Content-Type": "text/html; charset=utf-8"
            }
        )
    return {"detail": "Not Found"}


@app.get("/dashboard.html")
async def dashboard():
    """Sirve la página de dashboard."""
    dashboard_path = os.path.join(frontend_path, "dashboard.html")
    if os.path.exists(dashboard_path):
        return FileResponse(
            dashboard_path, 
            media_type="text/html; charset=utf-8",
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Content-Type": "text/html; charset=utf-8"
            }
        )
    return {"detail": "Not Found"}


@app.get("/api.js")
async def api_js():
    """Sirve el archivo api.js."""
    api_path = os.path.join(frontend_path, "api.js")
    if os.path.exists(api_path):
        return FileResponse(
            api_path, 
            media_type="application/javascript; charset=utf-8", 
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Content-Type": "application/javascript; charset=utf-8"
            }
        )
    return {"detail": "Not Found"}


@app.get("/health")
async def health():
    """Endpoint de salud para verificar que el backend está activo."""
    try:
        # Verificar conexión a base de datos
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT 1")
        
        return {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "database": "connected",
            "version": "1.0.0"
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return JSONResponse(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            content={
                "status": "unhealthy",
                "timestamp": datetime.now().isoformat(),
                "error": str(e)
            }
        )


@app.get("/api/status")
async def api_status():
    """Endpoint de status detallado para monitoring."""
    try:
        # Test DB
        db_healthy = False
        user_count = 0
        try:
            with get_db() as conn:
                result = conn.execute("SELECT COUNT(*) as count FROM users").fetchone()
                user_count = result["count"]
                db_healthy = True
        except Exception as db_error:
            logger.error(f"DB health check error: {db_error}")
        
        return {
            "status": "ok" if db_healthy else "degraded",
            "timestamp": datetime.now().isoformat(),
            "services": {
                "database": "healthy" if db_healthy else "unhealthy",
                "api": "healthy"
            },
            "stats": {
                "total_users": user_count
            }
        }
    except Exception as e:
        logger.error(f"Status check failed: {e}")
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"status": "error", "message": str(e)}
        )


@app.post("/login", response_class=HTMLResponse)
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    """Procesa el formulario de login y vuelve a renderizar la plantilla con
    un mensaje de éxito/fracaso.
    """
    # Autenticación de ejemplo (solo demo)
    authenticated = (username == "admin" and password == "secret")
    message = "Autenticado correctamente" if authenticated else "Credenciales inválidas"

    context = {
        "request": request,
        "username": username,
        "authenticated": authenticated,
        "message": message,
    }
    return templates.TemplateResponse("login.html", context)


@app.post("/api/login")
async def api_login(username: str = Form(...), password: str = Form(...)):
    """API JSON para login desde el frontend - verifica contra DB."""
    try:
        logger.info(f"Intento de login para usuario: {username}")
        with get_db() as conn:
            cursor = conn.cursor()
            query, params = sql("SELECT password_hash FROM users WHERE username = ?", (username,))
            cursor.execute(query, params)
            row = cursor.fetchone()
            
            if not row:
                logger.warning(f"Usuario no encontrado: {username}")
                return {"username": username, "authenticated": False, "message": "Credenciales inválidas"}
            
            # Verificar contraseña con bcrypt
            password_hash = row["password_hash"] if isinstance(row, dict) else row[0]
            if bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8')):
                logger.info(f"Login exitoso para usuario: {username}")
                return {"username": username, "authenticated": True, "message": "Autenticado correctamente"}
            else:
                logger.warning(f"Contraseña incorrecta para usuario: {username}")
                return {"username": username, "authenticated": False, "message": "Credenciales inválidas"}
    except Exception as e:
        logger.error(f"Error en login para {username}: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al procesar login"})


@app.post("/api/register")
async def api_register(
    email: str = Form(...),
    username: str = Form(...),
    birthdate: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
):
    """Registro: valida campos, crea usuario en DB con contraseña hasheada."""
    try:
        logger.info(f"Intento de registro para usuario: {username}")
        
        # Validaciones básicas
        if "@" not in email or "." not in email:
            raise HTTPException(status_code=400, detail={"message": "Correo inválido"})
        if len(username) < 3:
            raise HTTPException(status_code=400, detail={"message": "El usuario debe tener al menos 3 caracteres"})
        try:
            date.fromisoformat(birthdate)
        except Exception:
            raise HTTPException(status_code=400, detail={"message": "Fecha de nacimiento inválida (usa formato AAAA-MM-DD)"})
        if len(password) < 6:
            raise HTTPException(status_code=400, detail={"message": "La contraseña debe tener al menos 6 caracteres"})
        if password != confirm_password:
            raise HTTPException(status_code=400, detail={"message": "Las contraseñas no coinciden"})

        # Hashear contraseña con bcrypt
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        # Insertar usuario en la base de datos
        try:
            with get_db() as conn:
                cursor = conn.cursor()
                if USE_POSTGRES:
                    cursor.execute(
                        "INSERT INTO users (email, username, birthdate, password_hash) VALUES (%s, %s, %s, %s)",
                        (email, username, birthdate, password_hash)
                    )
                else:
                    cursor.execute(
                        "INSERT INTO users (email, username, birthdate, password_hash) VALUES (?, ?, ?, ?)",
                        (email, username, birthdate, password_hash)
                    )
                conn.commit()
            logger.info(f"Usuario registrado exitosamente: {username}")
        except Exception as e:
            logger.warning(f"Registro fallido - usuario o email duplicado: {username}")
            raise HTTPException(status_code=400, detail={"message": "El usuario o correo ya existe"})

        return {"success": True, "message": "Registro exitoso"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error inesperado en registro: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al procesar registro"})


# ----- Endpoints de Ajustes / Perfil -----
@app.get("/api/user")
async def get_user(username: str):
    """Obtiene datos públicos del usuario (email, username, birthdate, created_at)."""
    with get_db() as conn:
        cursor = conn.cursor()
        query, params = sql("SELECT email, username, birthdate, created_at FROM users WHERE username = ?", (username,))
        cursor.execute(query, params)
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
        email = row["email"] if isinstance(row, dict) else row[0]
        user = row["username"] if isinstance(row, dict) else row[1]
        birthdate = row["birthdate"] if isinstance(row, dict) else row[2]
        created = row["created_at"] if isinstance(row, dict) else row[3]
        return {"email": email, "username": user, "birthdate": birthdate, "created_at": created}


@app.post("/api/settings/update-email")
async def update_email(username: str = Form(...), email: str = Form(...)):
    """Actualiza el correo del usuario con validación básica y unicidad."""
    if "@" not in email or "." not in email:
        raise HTTPException(status_code=400, detail={"message": "Correo inválido"})
    with get_db() as conn:
        cursor = conn.cursor()
        # Verificar que usuario exista
        if USE_POSTGRES:
            cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
        else:
            cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
        # Intentar actualizar
        try:
            if USE_POSTGRES:
                cursor.execute("UPDATE users SET email = %s WHERE username = %s", (email, username))
            else:
                cursor.execute("UPDATE users SET email = ? WHERE username = ?", (email, username))
            conn.commit()
        except Exception:
            raise HTTPException(status_code=400, detail={"message": "El correo ya está en uso"})
    return {"success": True}


@app.post("/api/settings/update-username")
async def update_username(username: str = Form(...), new_username: str = Form(...)):
    """Actualiza el nombre de usuario (mínimo 3 caracteres y único)."""
    if len(new_username) < 3:
        raise HTTPException(status_code=400, detail={"message": "El usuario debe tener al menos 3 caracteres"})
    with get_db() as conn:
        cursor = conn.cursor()
        if USE_POSTGRES:
            cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
        else:
            cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
        try:
            if USE_POSTGRES:
                cursor.execute("UPDATE users SET username = %s WHERE username = %s", (new_username, username))
            else:
                cursor.execute("UPDATE users SET username = ? WHERE username = ?", (new_username, username))
            conn.commit()
        except Exception:
            raise HTTPException(status_code=400, detail={"message": "El usuario ya existe"})
    return {"success": True}


@app.post("/api/settings/update-password")
async def update_password(
    username: str = Form(...),
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
):
    """Actualiza la contraseña del usuario verificando la actual y validando la nueva."""
    if new_password != confirm_password:
        raise HTTPException(status_code=400, detail={"message": "Las contraseñas no coinciden"})
    # Validación de complejidad similar al front
    if len(new_password) <= 6 or not re.search(r"[a-z]", new_password) or not re.search(r"[A-Z]", new_password) or not re.search(r"\d", new_password):
        raise HTTPException(status_code=400, detail={"message": "La contraseña no cumple los criterios"})

    with get_db() as conn:
        cursor = conn.cursor()
        query, params = sql("SELECT password_hash FROM users WHERE username = ?", (username,))
        cursor.execute(query, params)
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
        # verificar contraseña actual
        password_hash = row["password_hash"] if isinstance(row, dict) else row[0]
        if not bcrypt.checkpw(current_password.encode('utf-8'), password_hash.encode('utf-8')):
            raise HTTPException(status_code=400, detail={"message": "La contraseña actual es incorrecta"})
        # actualizar
        new_hash = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        query, params = sql("UPDATE users SET password_hash = ? WHERE username = ?", (new_hash, username))
        cursor.execute(query, params)
        conn.commit()
    return {"success": True}


# ===============================================
# ENDPOINTS PARA CLIENTES
# ===============================================

@app.get("/api/clientes")
async def get_clientes(username: str):
    """Obtiene todos los clientes del usuario."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Obtener clientes
            cursor = conn.execute(
                "SELECT id, nombre, cedula, obra, estado, fecha, created_at FROM clientes WHERE user_id = ? ORDER BY created_at DESC",
                (user["id"],)
            )
            clientes = [dict(row) for row in cursor.fetchall()]
            return {"clientes": clientes}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al obtener clientes: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al obtener clientes"})


@app.post("/api/clientes")
async def create_cliente(
    username: str = Form(...),
    nombre: str = Form(...),
    cedula: str = Form(None),
    obra: str = Form(None),
    estado: str = Form("activo"),
    fecha: str = Form(None)
):
    """Crea un nuevo cliente."""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            # Obtener user_id
            cursor.execute(*sql("SELECT id FROM users WHERE username = ?", (username,)))
            user = cursor.fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Insertar cliente y obtener ID
            if USE_POSTGRES:
                cursor.execute(
                    "INSERT INTO clientes (user_id, nombre, cedula, obra, estado, fecha) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id",
                    (user["id"] if isinstance(user, dict) else user[0], nombre, cedula, obra, estado, fecha)
                )
                new_id = cursor.fetchone()[0]
            else:
                cursor.execute(
                    "INSERT INTO clientes (user_id, nombre, cedula, obra, estado, fecha) VALUES (?, ?, ?, ?, ?, ?)",
                    (user["id"] if isinstance(user, dict) else user[0], nombre, cedula, obra, estado, fecha)
                )
                new_id = cursor.lastrowid
            conn.commit()
            
            return {"success": True, "id": new_id}
    except Exception as e:
        logger.error(f"Error al crear cliente: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al crear cliente"})


@app.put("/api/clientes/{cliente_id}")
async def update_cliente(
    cliente_id: int,
    username: str = Form(...),
    nombre: str = Form(...),
    cedula: str = Form(None),
    obra: str = Form(None),
    estado: str = Form("activo"),
    fecha: str = Form(None)
):
    """Actualiza un cliente existente."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Verificar que el cliente pertenece al usuario
            cliente = conn.execute(
                "SELECT id FROM clientes WHERE id = ? AND user_id = ?",
                (cliente_id, user["id"])
            ).fetchone()
            
            if not cliente:
                raise HTTPException(status_code=404, detail={"message": "Cliente no encontrado"})
            
            # Actualizar cliente
            conn.execute(
                "UPDATE clientes SET nombre = ?, cedula = ?, obra = ?, estado = ?, fecha = ? WHERE id = ?",
                (nombre, cedula, obra, estado, fecha, cliente_id)
            )
            conn.commit()
            
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar cliente: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al actualizar cliente"})


@app.delete("/api/clientes/{cliente_id}")
async def delete_cliente(cliente_id: int, username: str):
    """Elimina un cliente."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Verificar y eliminar
            result = conn.execute(
                "DELETE FROM clientes WHERE id = ? AND user_id = ?",
                (cliente_id, user["id"])
            )
            conn.commit()
            
            if result.rowcount == 0:
                raise HTTPException(status_code=404, detail={"message": "Cliente no encontrado"})
            
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar cliente: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al eliminar cliente"})


# ===============================================
# ENDPOINTS PARA OBRAS
# ===============================================

@app.get("/api/obras")
async def get_obras(username: str):
    """Obtiene todas las obras del usuario."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Obtener obras
            cursor = conn.execute(
                "SELECT id, nombre, ubicacion, estado, created_at FROM obras WHERE user_id = ? ORDER BY created_at DESC",
                (user["id"],)
            )
            obras = [dict(row) for row in cursor.fetchall()]
            return {"obras": obras}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al obtener obras: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al obtener obras"})


@app.post("/api/obras")
async def create_obra(
    username: str = Form(...),
    nombre: str = Form(...),
    ubicacion: str = Form(None),
    estado: str = Form("activa")
):
    """Crea una nueva obra."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Insertar obra
            cursor = conn.execute(
                "INSERT INTO obras (user_id, nombre, ubicacion, estado) VALUES (?, ?, ?, ?)",
                (user["id"], nombre, ubicacion, estado)
            )
            conn.commit()
            
            return {"success": True, "id": cursor.lastrowid}
    except Exception as e:
        logger.error(f"Error al crear obra: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al crear obra"})


@app.put("/api/obras/{obra_id}")
async def update_obra(
    obra_id: int,
    username: str = Form(...),
    nombre: str = Form(...),
    ubicacion: str = Form(None),
    estado: str = Form("activa")
):
    """Actualiza una obra existente."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Verificar que la obra pertenece al usuario
            obra = conn.execute(
                "SELECT id FROM obras WHERE id = ? AND user_id = ?",
                (obra_id, user["id"])
            ).fetchone()
            
            if not obra:
                raise HTTPException(status_code=404, detail={"message": "Obra no encontrada"})
            
            # Actualizar obra
            conn.execute(
                "UPDATE obras SET nombre = ?, ubicacion = ?, estado = ? WHERE id = ?",
                (nombre, ubicacion, estado, obra_id)
            )
            conn.commit()
            
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar obra: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al actualizar obra"})


@app.delete("/api/obras/{obra_id}")
async def delete_obra(obra_id: int, username: str):
    """Elimina una obra."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Verificar y eliminar
            result = conn.execute(
                "DELETE FROM obras WHERE id = ? AND user_id = ?",
                (obra_id, user["id"])
            )
            conn.commit()
            
            if result.rowcount == 0:
                raise HTTPException(status_code=404, detail={"message": "Obra no encontrada"})
            
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar obra: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al eliminar obra"})


# ===============================================
# ENDPOINTS PARA PRODUCTOS
# ===============================================

@app.get("/api/productos")
async def get_productos(username: str):
    """Obtiene todos los productos del usuario."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Obtener productos
            cursor = conn.execute(
                "SELECT id, nombre, precio, created_at FROM productos WHERE user_id = ? ORDER BY created_at DESC",
                (user["id"],)
            )
            productos = [dict(row) for row in cursor.fetchall()]
            return {"productos": productos}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al obtener productos: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al obtener productos"})


@app.post("/api/productos")
async def create_producto(
    username: str = Form(...),
    nombre: str = Form(...),
    precio: float = Form(0)
):
    """Crea un nuevo producto."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Insertar producto
            cursor = conn.execute(
                "INSERT INTO productos (user_id, nombre, precio) VALUES (?, ?, ?)",
                (user["id"], nombre, precio)
            )
            conn.commit()
            
            return {"success": True, "id": cursor.lastrowid}
    except Exception as e:
        logger.error(f"Error al crear producto: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al crear producto"})


@app.put("/api/productos/{producto_id}")
async def update_producto(
    producto_id: int,
    username: str = Form(...),
    nombre: str = Form(...),
    precio: float = Form(0)
):
    """Actualiza un producto existente."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Verificar que el producto pertenece al usuario
            producto = conn.execute(
                "SELECT id FROM productos WHERE id = ? AND user_id = ?",
                (producto_id, user["id"])
            ).fetchone()
            
            if not producto:
                raise HTTPException(status_code=404, detail={"message": "Producto no encontrado"})
            
            # Actualizar producto
            conn.execute(
                "UPDATE productos SET nombre = ?, precio = ? WHERE id = ?",
                (nombre, precio, producto_id)
            )
            conn.commit()
            
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar producto: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al actualizar producto"})


@app.delete("/api/productos/{producto_id}")
async def delete_producto(producto_id: int, username: str):
    """Elimina un producto."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Verificar y eliminar
            result = conn.execute(
                "DELETE FROM productos WHERE id = ? AND user_id = ?",
                (producto_id, user["id"])
            )
            conn.commit()
            
            if result.rowcount == 0:
                raise HTTPException(status_code=404, detail={"message": "Producto no encontrado"})
            
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar producto: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al eliminar producto"})


# ===============================================
# ENDPOINTS PARA REGISTROS
# ===============================================

@app.get("/api/registros")
async def get_registros(username: str, obra: str = None, fecha_inicio: str = None, fecha_fin: str = None):
    """Obtiene todos los registros del usuario con filtros opcionales."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Construir query con filtros
            query = "SELECT id, fecha, obra, totalCantidad, totalCobrar, totalPagado, status, clientesAdicionales, detalles, created_at FROM registros WHERE user_id = ?"
            params = [user["id"]]
            
            if obra:
                query += " AND obra = ?"
                params.append(obra)
            
            if fecha_inicio:
                query += " AND fecha >= ?"
                params.append(fecha_inicio)
            
            if fecha_fin:
                query += " AND fecha <= ?"
                params.append(fecha_fin)
            
            query += " ORDER BY fecha DESC, created_at DESC"
            
            # Obtener registros
            cursor = conn.execute(query, params)
            registros = [dict(row) for row in cursor.fetchall()]
            
            # Parsear JSON fields
            for registro in registros:
                if registro.get('clientesAdicionales'):
                    try:
                        import json
                        registro['clientesAdicionales'] = json.loads(registro['clientesAdicionales'])
                    except:
                        registro['clientesAdicionales'] = []
                else:
                    registro['clientesAdicionales'] = []
                    
                if registro.get('detalles'):
                    try:
                        import json
                        registro['detalles'] = json.loads(registro['detalles'])
                    except:
                        registro['detalles'] = []
                else:
                    registro['detalles'] = []
            
            return {"registros": registros}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al obtener registros: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al obtener registros"})


@app.post("/api/registros")
async def create_registro(request: Request):
    """Crea un nuevo registro."""
    try:
        import json
        body = await request.json()
        
        username = body.get('username')
        fecha = body.get('fecha')
        obra = body.get('obra')
        totalCantidad = body.get('totalCantidad', 0)
        totalCobrar = body.get('totalCobrar', 0)
        totalPagado = body.get('totalPagado', 0)
        status = body.get('status', 'pendiente')
        clientesAdicionales = body.get('clientesAdicionales', [])
        detalles = body.get('detalles', [])
        
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Convertir listas a JSON
            clientesAdicionales_json = json.dumps(clientesAdicionales) if clientesAdicionales else None
            detalles_json = json.dumps(detalles) if detalles else None
            
            # Insertar registro
            cursor = conn.execute(
                """INSERT INTO registros 
                   (user_id, fecha, obra, totalCantidad, totalCobrar, totalPagado, status, clientesAdicionales, detalles) 
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (user["id"], fecha, obra, totalCantidad, totalCobrar, totalPagado, status, 
                 clientesAdicionales_json, detalles_json)
            )
            conn.commit()
            
            return {"success": True, "id": cursor.lastrowid}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al crear registro: {e}")
        raise HTTPException(status_code=500, detail={"message": f"Error al crear registro: {str(e)}"})


@app.put("/api/registros/{registro_id}")
async def update_registro(registro_id: int, request: Request):
    """Actualiza un registro existente."""
    try:
        import json
        body = await request.json()
        
        username = body.get('username')
        fecha = body.get('fecha')
        obra = body.get('obra')
        totalCantidad = body.get('totalCantidad', 0)
        totalCobrar = body.get('totalCobrar', 0)
        totalPagado = body.get('totalPagado', 0)
        status = body.get('status', 'pendiente')
        clientesAdicionales = body.get('clientesAdicionales', [])
        detalles = body.get('detalles', [])
        
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Verificar que el registro pertenece al usuario
            registro = conn.execute(
                "SELECT id FROM registros WHERE id = ? AND user_id = ?",
                (registro_id, user["id"])
            ).fetchone()
            
            if not registro:
                raise HTTPException(status_code=404, detail={"message": "Registro no encontrado"})
            
            # Convertir listas a JSON
            clientesAdicionales_json = json.dumps(clientesAdicionales) if clientesAdicionales else None
            detalles_json = json.dumps(detalles) if detalles else None
            
            # Actualizar registro
            conn.execute(
                """UPDATE registros 
                   SET fecha = ?, obra = ?, totalCantidad = ?, totalCobrar = ?, 
                       totalPagado = ?, status = ?, clientesAdicionales = ?, detalles = ?
                   WHERE id = ?""",
                (fecha, obra, totalCantidad, totalCobrar, totalPagado, status, 
                 clientesAdicionales_json, detalles_json, registro_id)
            )
            conn.commit()
            
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar registro: {e}")
        raise HTTPException(status_code=500, detail={"message": f"Error al actualizar registro: {str(e)}"})


@app.delete("/api/registros/{registro_id}")
async def delete_registro(registro_id: int, username: str):
    """Elimina un registro."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Verificar y eliminar
            result = conn.execute(
                "DELETE FROM registros WHERE id = ? AND user_id = ?",
                (registro_id, user["id"])
            )
            conn.commit()
            
            if result.rowcount == 0:
                raise HTTPException(status_code=404, detail={"message": "Registro no encontrado"})
            
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar registro: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al eliminar registro"})


# ===============================================
# ENDPOINT PARA REPORTES/ESTADÍSTICAS
# ===============================================

@app.get("/api/reportes")
async def get_reportes(username: str, obra: str = None, fecha_inicio: str = None, fecha_fin: str = None):
    """Genera estadísticas y reportes basados en los registros."""
    try:
        with get_db() as conn:
            # Obtener user_id
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            # Construir query con filtros
            query = "SELECT fecha, obra, totalCantidad, totalCobrar, totalPagado, status FROM registros WHERE user_id = ?"
            params = [user["id"]]
            
            if obra:
                query += " AND obra = ?"
                params.append(obra)
            
            if fecha_inicio:
                query += " AND fecha >= ?"
                params.append(fecha_inicio)
            
            if fecha_fin:
                query += " AND fecha <= ?"
                params.append(fecha_fin)
            
            query += " ORDER BY fecha DESC"
            
            # Obtener registros
            cursor = conn.execute(query, params)
            registros = [dict(row) for row in cursor.fetchall()]
            
            # Calcular estadísticas
            total_cobrar = sum(r['totalCobrar'] or 0 for r in registros)
            total_cobrado = sum(r['totalPagado'] or 0 for r in registros)
            total_pendiente = total_cobrar - total_cobrado
            total_cantidad = sum(r['totalCantidad'] or 0 for r in registros)
            
            # Agrupar por obra
            por_obra = {}
            for r in registros:
                obra_nombre = r['obra'] or 'Sin obra'
                if obra_nombre not in por_obra:
                    por_obra[obra_nombre] = {
                        'totalCobrar': 0,
                        'totalCobrado': 0,
                        'totalPendiente': 0,
                        'totalCantidad': 0
                    }
                por_obra[obra_nombre]['totalCobrar'] += r['totalCobrar'] or 0
                por_obra[obra_nombre]['totalCobrado'] += r['totalPagado'] or 0
                por_obra[obra_nombre]['totalPendiente'] += (r['totalCobrar'] or 0) - (r['totalPagado'] or 0)
                por_obra[obra_nombre]['totalCantidad'] += r['totalCantidad'] or 0
            
            # Agrupar por fecha
            por_fecha = {}
            for r in registros:
                fecha = r['fecha'] or 'Sin fecha'
                if fecha not in por_fecha:
                    por_fecha[fecha] = {
                        'totalCobrar': 0,
                        'totalCobrado': 0,
                        'totalPendiente': 0,
                        'totalCantidad': 0
                    }
                por_fecha[fecha]['totalCobrar'] += r['totalCobrar'] or 0
                por_fecha[fecha]['totalCobrado'] += r['totalPagado'] or 0
                por_fecha[fecha]['totalPendiente'] += (r['totalCobrar'] or 0) - (r['totalPagado'] or 0)
                por_fecha[fecha]['totalCantidad'] += r['totalCantidad'] or 0
            
            return {
                "totales": {
                    "totalCobrar": total_cobrar,
                    "totalCobrado": total_cobrado,
                    "totalPendiente": total_pendiente,
                    "totalCantidad": total_cantidad,
                    "totalRegistros": len(registros)
                },
                "porObra": por_obra,
                "porFecha": por_fecha,
                "registros": registros
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al generar reportes: {e}")
        raise HTTPException(status_code=500, detail={"message": "Error al generar reportes"})


@app.post("/api/import-backup")
async def import_backup(request: Request):
    """Importa clientes, obras, productos, registros en bulk desde un JSON."""
    try:
        import json
        body = await request.json()
        
        username = body.get('username')
        clientes_data = body.get('clientes', [])
        obras_data = body.get('obras', [])
        productos_data = body.get('productos', [])
        registros_data = body.get('registros', [])
        
        with get_db() as conn:
            user = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail={"message": "Usuario no encontrado"})
            
            user_id = user["id"]
            counts = {"clientes": 0, "obras": 0, "productos": 0, "registros": 0}
            
            for c in clientes_data:
                try:
                    conn.execute(
                        "INSERT INTO clientes (user_id, nombre, cedula, obra, estado, fecha) VALUES (?, ?, ?, ?, ?, ?)",
                        (user_id, c.get("nombre", ""), c.get("cedula"), c.get("obra"), c.get("estado", "activo"), c.get("fecha"))
                    )
                    counts["clientes"] += 1
                except Exception as e:
                    logger.warning(f"Error importing cliente: {e}")
            
            for o in obras_data:
                try:
                    conn.execute(
                        "INSERT INTO obras (user_id, nombre, ubicacion, estado) VALUES (?, ?, ?, ?)",
                        (user_id, o.get("nombre", ""), o.get("ubicacion"), o.get("estado", "activa"))
                    )
                    counts["obras"] += 1
                except Exception as e:
                    logger.warning(f"Error importing obra: {e}")
            
            for p in productos_data:
                try:
                    conn.execute(
                        "INSERT INTO productos (user_id, nombre, precio) VALUES (?, ?, ?)",
                        (user_id, p.get("nombre", ""), float(p.get("precio", 0)))
                    )
                    counts["productos"] += 1
                except Exception as e:
                    logger.warning(f"Error importing producto: {e}")
            
            for r in registros_data:
                try:
                    detalles = r.get("items") or r.get("detalles") or []
                    detalles_json = json.dumps(detalles) if detalles else None
                    
                    adicionales = []
                    for it in detalles:
                        if str(it.get("tipo", "")).lower() == "adicional":
                            adicionales.append({"cliente": it.get("cliente"), "valor": it.get("costo", it.get("precio", 0))})
                    adicionales_json = json.dumps(adicionales) if adicionales else None
                    
                    conn.execute(
                        """INSERT INTO registros 
                           (user_id, fecha, obra, totalCantidad, totalCobrar, totalPagado, status, clientesAdicionales, detalles)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (user_id, r.get("fecha"), r.get("obra"), r.get("totalCantidad", 0), 
                         r.get("totalCobrar", 0), r.get("totalPagado", 0), r.get("status", "pendiente"),
                         adicionales_json, detalles_json)
                    )
                    counts["registros"] += 1
                except Exception as e:
                    logger.warning(f"Error importing registro: {e}")
            
            conn.commit()
            return {"success": True, "imported": counts}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en import-backup: {e}")
        raise HTTPException(status_code=500, detail={"message": f"Error al importar: {str(e)}"})


# ----- Endpoints de Administración (con clave secreta) -----
ADMIN_SECRET = os.getenv("ADMIN_SECRET", "admin_secret_key_2026")

@app.post("/api/admin/verify-password")
async def admin_verify_password(
    username: str = Form(...),
    password: str = Form(...),
    admin_secret: str = Form(...)
):
    """Endpoint de administración para verificar si una contraseña es correcta."""
    if admin_secret != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail={"message": "Acceso denegado"})
    
    try:
        with get_db() as conn:
            cursor = conn.execute(
                "SELECT password_hash FROM users WHERE username = ?",
                (username,)
            )
            row = cursor.fetchone()
            
            if not row:
                return {"success": False, "message": f"Usuario '{username}' no encontrado"}
            
            password_hash = row["password_hash"]
            is_valid = bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))
            
            return {
                "success": True,
                "username": username,
                "password_valid": is_valid,
                "message": "Contraseña correcta" if is_valid else "Contraseña incorrecta"
            }
    except Exception as e:
        logger.error(f"Error verificando contraseña: {e}")
        raise HTTPException(status_code=500, detail={"message": str(e)})


@app.post("/api/admin/reset-password")
async def admin_reset_password(
    username: str = Form(...),
    new_password: str = Form(...),
    admin_secret: str = Form(...)
):
    """Endpoint de administración para resetear contraseña de un usuario."""
    if admin_secret != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail={"message": "Acceso denegado"})
    
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail={"message": "La contraseña debe tener al menos 6 caracteres"})
    
    try:
        with get_db() as conn:
            # Verificar que usuario existe
            cursor = conn.execute("SELECT id FROM users WHERE username = ?", (username,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail={"message": f"Usuario '{username}' no encontrado"})
            
            # Hashear nueva contraseña
            password_hash = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # Actualizar contraseña
            conn.execute("UPDATE users SET password_hash = ? WHERE username = ?", (password_hash, username))
            conn.commit()
            
            logger.info(f"Contraseña reseteada exitosamente para usuario: {username}")
            return {
                "success": True,
                "message": f"Contraseña actualizada para '{username}'",
                "username": username
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error reseteando contraseña: {e}")
        raise HTTPException(status_code=500, detail={"message": str(e)})


@app.get("/api/admin/list-users")
async def admin_list_users(admin_secret: str):
    """Lista todos los usuarios (solo email y username, sin contraseñas)."""
    if admin_secret != ADMIN_SECRET:
        raise HTTPException(status_code=403, detail={"message": "Acceso denegado"})
    
    try:
        with get_db() as conn:
            cursor = conn.execute(
                "SELECT id, email, username, birthdate, created_at FROM users ORDER BY created_at DESC"
            )
            users = []
            for row in cursor.fetchall():
                users.append({
                    "id": row["id"],
                    "email": row["email"],
                    "username": row["username"],
                    "birthdate": row["birthdate"],
                    "created_at": row["created_at"]
                })
            
            return {"success": True, "users": users, "count": len(users)}
    except Exception as e:
        logger.error(f"Error listando usuarios: {e}")
        raise HTTPException(status_code=500, detail={"message": str(e)})


# Exportar Excel con estilos profesionales completos
@app.post("/api/reportes/export-excel")
async def export_reportes_excel(request: Request):
    """Genera archivo Excel con estilos completos usando openpyxl.
    
    Formato garantizado:
    - Usuario: Arial 12, Negrita, Centrado, Celda combinada
    - Título: Arial 18, Negrita, Centrado, Celda combinada
    - Fechas: Arial 12, Negrita, Centrado, Celda combinada
    - Encabezados de columnas: Arial 11, Negrita, Fondo gris
    - Datos: Arial 10
    - Totales: Negrita, Fondo gris
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import Response
        import io

        data = await request.json()
        rows = data.get('rows', [])
        headers = data.get('headers', [])
        username = data.get('username', 'Usuario')
        title = data.get('title', 'Reporte')
        date_range = data.get('date_range', '')
        currency_cols = data.get('currency_cols', [])
        mode = data.get('mode', 'general')
        totals = data.get('totals', None)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Reportes'

        # Definir estilos profesionales
        font_usuario = Font(name='Arial', bold=True, size=12)
        font_titulo = Font(name='Arial', bold=True, size=18)
        font_fechas = Font(name='Arial', bold=True, size=12)
        font_col_header = Font(name='Arial', bold=True, size=11)
        font_data = Font(name='Arial', size=10)
        font_totals = Font(name='Arial', bold=True, size=10)
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        
        fill_gray = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        fill_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Fila 1: Usuario (combinar celdas)
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        cell = ws['A1']
        cell.value = username
        cell.font = font_usuario
        cell.alignment = align_center

        # Fila 2: Título (combinar celdas)
        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        cell = ws['A2']
        cell.value = title
        cell.font = font_titulo
        cell.alignment = align_center

        # Fila 3: Rango de fechas (combinar celdas)
        ws.merge_cells(f'A3:{get_column_letter(len(headers))}3')
        cell = ws['A3']
        cell.value = date_range
        cell.font = font_fechas
        cell.alignment = align_center

        # Fila 4: vacía (espacio)
        
        # Fila 5: Encabezados de columnas
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = font_col_header
            cell.alignment = align_center
            cell.fill = fill_header
            cell.border = thin_border

        # Filas 6+: Datos
        for r_idx, row in enumerate(rows, start=6):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = font_data
                cell.border = thin_border
                
                # Alineación según tipo de dato
                if (c_idx - 1) in currency_cols:
                    cell.alignment = align_right
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                elif headers[c_idx - 1] in ['Estado']:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        # Fila de totales (si existe)
        if totals:
            totals_row = len(rows) + 6
            for c_idx, value in enumerate(totals, start=1):
                cell = ws.cell(row=totals_row, column=c_idx, value=value)
                cell.font = font_totals
                cell.fill = fill_gray
                cell.border = thin_border
                
                if (c_idx - 1) in currency_cols and isinstance(value, (int, float)):
                    cell.alignment = align_right
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = align_right if c_idx > 1 else align_left

        # Ajustar anchos de columna por modo
        col_widths_map = {
            'general': [16, 22, 12, 16, 16, 16, 12],
            'detallado': [16, 22, 16, 22, 12, 16, 16, 16, 12],
            'diario': [22, 22] + [14] * max(0, (len(headers) - 3)) + [16]
        }
        widths = col_widths_map.get(mode, [16] * len(headers))
        for c_idx, width in enumerate(widths, start=1):
            if c_idx <= len(headers):
                ws.column_dimensions[get_column_letter(c_idx)].width = width

        # Generar archivo
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"reportes_{mode}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )
    except Exception as e:
        logger.error(f"Error en export_reportes_excel: {e}")
        raise HTTPException(status_code=500, detail={"message": str(e)})


if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000, reload=True)
